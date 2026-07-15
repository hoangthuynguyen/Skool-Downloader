#!/usr/bin/env python3
"""
Cập nhật file Excel danh sách communities Skool (format người dùng).

Đọc workbook cũ (Latest), cào lại Discovery (Playwright) theo:
  - mọi ngôn ngữ
  - mọi chủ đề
  - free / paid
  - category × language (ngôn ngữ lớn)
Gộp theo slug, cập nhật Members/Cost, thêm community mới.
Ghi sheet Latest + sheet theo ngày (YYYY-MM).
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urlencode

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

DEFAULT_SRC = Path.home() / "Downloads" / "Skool 14-05-2026.xlsx"
HEADERS = [
    "#",
    "Name",
    "Tagline",
    "Members",
    "Cost / Month",
    "Total / Month",
    "Language",
    "Category",
    "Notes",
    "URL",
]

LANG_MAP = {
    "english": "English",
    "german": "German",
    "spanish": "Spanish",
    "french": "French",
    "chinese": "Chinese",
    "italian": "Italian",
    "dutch": "Dutch",
    "vietnamese": "Vietnamese",
    "arabic": "Arabic",
    "hebrew": "Hebrew",
    "danish": "Danish",
    "romanian": "Romanian",
    "turkish": "Turkish",
    "polish": "Polish",
    "czech": "Czech",
    "hungarian": "Hungarian",
    "swedish": "Swedish",
    "portuguese": "Portuguese",
    "bulgarian": "Bulgarian",
    "norwegian": "Norwegian",
    "finnish": "Finnish",
    "croatian": "Croatian",
    "latvian": "Latvian",
    "slovak": "Slovak",
    "serbian": "Serbian",
    "mongolian": "Mongolian",
    "haitian": "Haitian",
    "thai": "Thai",
    "slovenian": "Slovenian",
    "russian": "Russian",
    "lithuanian": "Lithuanian",
    "amharic": "Amharic",
    "malay": "Malay",
    "estonian": "Estonian",
    "greek": "Greek",
    "ukrainian": "Ukrainian",
    "swahili": "Swahili",
    "japanese": "Japanese",
    "filipino": "Filipino",
    "persian": "Persian",
    "welsh": "Welsh",
    "korean": "Korean",
    "cantonese": "Cantonese",
    "indonesian": "Indonesian",
    "latin": "Latin",
    "bengali": "Bengali",
    "catalan": "Catalan",
    "hindi": "Hindi",
}

# lang keys đủ lớn để nhân category
BIG_LANGS = ("english", "german", "spanish", "french", "portuguese", "chinese", "italian")


def log(msg: str) -> None:
    print(msg, flush=True)


def slug_from_url(url: str) -> str:
    u = (url or "").strip().rstrip("/")
    u = u.replace("https://www.skool.com/", "").replace("http://www.skool.com/", "")
    u = u.replace("https://skool.com/", "")
    if u.endswith("/about"):
        u = u[: -len("/about")]
    return u.split("/")[0].strip()


def clean_category(name: str) -> str:
    s = re.sub(
        "[\U0001F000-\U0001FAFF\U00002600-\U000027BF\U0001F1E6-\U0001F1FF"
        "\u2190-\u21FF\u2B00-\u2BFF\u2300-\u23FF\uFE0F\u200D]",
        "",
        name or "",
    )
    return re.sub(r"\s+", " ", s).strip()


def parse_cost_display(meta: dict) -> Tuple[float, str]:
    """Trả (cost_per_month_float, notes). Free → 0, 'Free'."""
    model = meta.get("membershipModel")
    try:
        model_i = int(model) if model is not None else None
    except (TypeError, ValueError):
        model_i = None

    def parse_json_price(raw, force_interval=None):
        if not raw:
            return None
        try:
            d = json.loads(raw) if isinstance(raw, str) else raw
            if not isinstance(d, dict):
                return None
            amount = float(d.get("amount") or 0) / 100.0
            iv = force_interval or (d.get("recurring_interval") or "month")
            iv = str(iv).lower()
            if amount <= 0:
                return 0.0, "Free"
            if iv in ("year", "yearly", "annual", "yr"):
                return round(amount / 12.0, 2), f"Yearly ${amount:,.0f} ÷12"
            if iv in ("one_time", "onetime", "once"):
                return 0.0, f"One-time ${amount:,.2f}"
            return amount, ""
        except Exception:
            return None

    for key, iv in (
        ("displayPrice", None),
        ("currentMBp", "month"),
        ("currentABp", "year"),
        ("currentOtBp", "one_time"),
    ):
        hit = parse_json_price(meta.get(key), iv)
        if hit is not None:
            return hit

    # Free / freemium models without price
    if model_i in (0, 1, 3) or model_i is None:
        return 0.0, "Free"
    return 0.0, "Free"


def normalize_cost_raw(s) -> float:
    s = str(s or "")
    # "$0 (Free)" / "Free" → 0; keep real dollar amounts
    if s.strip() in ("$0", "$0 (Free)", "", "None", "0", "0.0", "Free"):
        return 0.0
    if re.fullmatch(r"(?i)free", s.strip()):
        return 0.0
    m = re.search(r"\$([\d,]+(?:\.\d+)?)", s)
    if m:
        return float(m.group(1).replace(",", ""))
    # bare number
    m2 = re.search(r"([\d]+(?:\.\d+)?)", s.replace(",", ""))
    if m2 and "free" not in s.lower():
        try:
            return float(m2.group(1))
        except ValueError:
            return 0.0
    return 0.0


def normalize_notes(cost: float, notes: Any) -> str:
    """
    Chuẩn hoá Notes cho khớp Cost:
      - cost > 0 → không được ghi 'Free'
      - cost <= 0 + không có ghi chú đặc biệt → 'Free'
      - giữ Yearly / One-time / Missing on Skool / not in latest scrape
    """
    try:
        c = float(cost or 0)
    except (TypeError, ValueError):
        c = 0.0
    n = str(notes or "").strip()
    n = re.sub(r"\s*\|\s*", " | ", n)
    n = re.sub(r"\s+", " ", n).strip(" |")

    # Tách flag stale (nếu có) rồi gắn lại sau
    stale = False
    if "not in latest scrape" in n.lower():
        stale = True
        n = re.sub(r"(?i)\s*\|\s*not in latest scrape", "", n)
        n = re.sub(r"(?i)not in latest scrape\s*\|\s*", "", n)
        n = re.sub(r"(?i)not in latest scrape", "", n).strip(" |")

    special_keep = (
        n.startswith("Yearly")
        or n.startswith("One-time")
        or n.startswith("Missing")
        or n.lower().startswith("paid")
    )

    if c > 0:
        # Paid: bỏ nhãn Free sai; giữ Yearly/One-time nếu có
        if n.lower() in ("free", "free trial", "freemium", ""):
            n = ""
        elif "free" == n.lower().split("|")[0].strip():
            n = ""
    else:
        # Free / one-time (cost 0): nếu notes rỗng → Free (trừ special)
        if not n and not special_keep:
            n = "Free"
        elif n.lower() in ("",):
            n = "Free"

    if stale:
        n = (n + " | not in latest scrape").strip(" |")
    return n


def parse_group(entry: dict, lang_tag: Optional[str] = None, cat_name: str = "") -> Optional[dict]:
    g = entry.get("group", entry) if isinstance(entry, dict) else None
    if not isinstance(g, dict):
        return None
    meta = g.get("metadata") or {}
    if not isinstance(meta, dict):
        meta = {}
    slug = (g.get("name") or "").strip()
    if not slug:
        return None
    name = (meta.get("displayName") or slug).strip()
    tagline = re.sub(r"\s+", " ", (meta.get("description") or "").strip())[:300]
    try:
        members = int(meta.get("totalMembers") or 0)
    except (TypeError, ValueError):
        members = 0
    cost, notes = parse_cost_display(meta)
    notes = normalize_notes(cost, notes)
    if lang_tag:
        lang = LANG_MAP.get(lang_tag.lower(), lang_tag.capitalize())
    else:
        lang = ""
    cat = clean_category(cat_name) if cat_name else ""
    return {
        "slug": slug,
        "name": name,
        "tagline": tagline,
        "members": members,
        "cost": cost,
        "notes": notes,
        "lang": lang,
        "category": cat,
        "url": f"https://www.skool.com/{slug}/about",
        "refreshed": True,
    }


def load_existing(path: Path) -> Dict[str, dict]:
    seen: Dict[str, dict] = {}
    if not path.exists():
        log(f"⚠ Không thấy file cũ: {path}")
        return seen
    wb = load_workbook(path, data_only=True)
    # ưu tiên sheet Latest, rồi 2026-05, rồi sheet đầu có header Name
    sheet = None
    for name in ("Latest", "2026-05", "Skool Communities"):
        if name in wb.sheetnames:
            sheet = wb[name]
            break
    if sheet is None:
        for name in wb.sheetnames:
            ws = wb[name]
            row1 = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            if "Name" in row1 and "URL" in row1:
                sheet = ws
                break
    if sheet is None:
        log("⚠ Không tìm thấy sheet dữ liệu trong Excel cũ")
        wb.close()
        return seen

    headers = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    # map alias Tagline column
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        d = dict(zip(headers, row))
        url = str(d.get("URL") or "").strip()
        name = str(d.get("Name") or "").strip()
        slug = slug_from_url(url) if url else ""
        if not slug and name:
            slug = re.sub(r"\s+", "-", name.lower())[:80]
        if not slug:
            continue
        members_raw = d.get("Members", 0) or 0
        try:
            members = int(float(str(members_raw).replace(",", "")))
        except Exception:
            members = 0
        tagline = d.get("Tagline")
        if tagline is None:
            tagline = d.get("Make more money, help more people.") or ""
        cost = normalize_cost_raw(d.get("Cost / Month"))
        notes = normalize_notes(cost, d.get("Notes"))
        seen[slug] = {
            "slug": slug,
            "name": name,
            "tagline": str(tagline or ""),
            "members": members,
            "cost": cost,
            "notes": notes,
            "lang": str(d.get("Language") or "English"),
            "category": str(d.get("Category") or ""),
            "url": url or f"https://www.skool.com/{slug}/about",
            "refreshed": False,
        }
    wb.close()
    log(f"📂 Loaded {len(seen):,} communities từ {path.name}")
    return seen


def merge_item(store: Dict[str, dict], item: dict) -> str:
    """Merge scraped item. Return 'new' | 'updated' | 'same'."""
    slug = item["slug"]
    old = store.get(slug)
    if not old:
        item = dict(item)
        item["notes"] = normalize_notes(item.get("cost") or 0, item.get("notes"))
        store[slug] = item
        return "new"
    # update live fields
    changed = False
    # Text fields: only overwrite when new value non-empty
    for k in ("name", "tagline", "url"):
        v = item.get(k)
        if v not in (None, "") and v != old.get(k):
            old[k] = v
            changed = True
    # Members: take newer positive/zero always when provided
    if item.get("members") is not None:
        try:
            m = int(item["members"])
            if m != int(old.get("members") or 0):
                old["members"] = m
                changed = True
        except (TypeError, ValueError):
            pass
    # Cost + notes always move together (empty notes for paid must clear stale "Free")
    cost_in = item.get("cost")
    if cost_in is not None:
        try:
            new_cost = float(cost_in)
        except (TypeError, ValueError):
            new_cost = float(old.get("cost") or 0)
        old_cost = float(old.get("cost") or 0)
        # Prefer scraped notes when key present (incl. ""), else keep old notes
        raw_notes = item["notes"] if "notes" in item else old.get("notes")
        new_notes = normalize_notes(new_cost, raw_notes)
        if abs(new_cost - old_cost) > 1e-9 or new_notes != (old.get("notes") or ""):
            old["cost"] = new_cost
            old["notes"] = new_notes
            changed = True
    elif "notes" in item:
        fixed = normalize_notes(old.get("cost") or 0, item.get("notes"))
        if fixed != (old.get("notes") or ""):
            old["notes"] = fixed
            changed = True
    else:
        sanitized = normalize_notes(old.get("cost") or 0, old.get("notes"))
        if sanitized != (old.get("notes") or ""):
            old["notes"] = sanitized
            changed = True

    if item.get("lang"):
        if old.get("lang") != item["lang"]:
            old["lang"] = item["lang"]
            changed = True
    if item.get("category") and not old.get("category"):
        old["category"] = item["category"]
        changed = True
    elif item.get("category") and item["category"] != old.get("category"):
        # prefer non-empty newer category when refreshed
        old["category"] = item["category"]
        changed = True
    old["refreshed"] = True
    return "updated" if changed else "same"


def save_excel(data: List[dict], path: Path, sheet_date: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # remove default
    default = wb.active
    wb.remove(default)

    def write_sheet(ws, rows):
        hfill = PatternFill("solid", fgColor="1E40AF")
        hfont = Font(bold=True, color="FFFFFF", size=11, name="Arial")
        alt = PatternFill("solid", fgColor="EFF6FF")
        paid = PatternFill("solid", fgColor="D1FAE5")
        stale = PatternFill("solid", fgColor="FEF3C7")
        body = Font(name="Arial", size=10)
        for col, h in enumerate(HEADERS, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hfill
            c.font = hfont
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 32
        for i, row in enumerate(rows, 1):
            r = i + 1
            try:
                cost = float(row.get("cost") or 0)
            except (TypeError, ValueError):
                cost = 0.0
            try:
                members = int(row.get("members") or 0)
            except (TypeError, ValueError):
                members = 0
            total = int(members * cost) if cost > 0 else 0
            notes = normalize_notes(cost, row.get("notes"))
            if not row.get("refreshed") and "not in latest scrape" not in notes.lower():
                notes = (notes + " | not in latest scrape").strip(" |")
            vals = [
                i,
                row.get("name") or "",
                row.get("tagline") or "",
                members,
                # Free = $0 (số $0 rõ ràng để tính tổng doanh thu = 0)
                f"${cost:,.2f}" if cost > 0 else "$0",
                f"${total:,.2f}" if total > 0 else "$0",
                row.get("lang") or "",
                row.get("category") or "",
                notes,
                row.get("url") or "",
            ]
            if not row.get("refreshed"):
                fill = stale
            elif cost > 0:
                fill = paid
            elif i % 2 == 0:
                fill = alt
            else:
                fill = None
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=col, value=v)
                cell.font = body
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if fill:
                    cell.fill = fill
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="right", vertical="top")
        widths = [5, 34, 55, 12, 14, 16, 14, 16, 28, 48]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
        ws.freeze_panes = "A2"
        if rows:
            ws.auto_filter.ref = f"A1:J{len(rows)+1}"

    sorted_rows = sorted(
        data, key=lambda x: (int(x.get("members") or 0), x.get("name") or ""), reverse=True
    )
    # Latest
    ws_latest = wb.create_sheet("Latest", 0)
    write_sheet(ws_latest, sorted_rows)
    # dated
    ws_dated = wb.create_sheet(sheet_date, 1)
    write_sheet(ws_dated, sorted_rows)

    # Meta sheet
    ws_m = wb.create_sheet("Meta", 2)
    ws_m["A1"] = "Updated at"
    ws_m["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws_m["A2"] = "Total communities"
    ws_m["B2"] = len(sorted_rows)
    ws_m["A3"] = "Refreshed in this scrape"
    ws_m["B3"] = sum(1 for r in sorted_rows if r.get("refreshed"))
    ws_m["A4"] = "Kept from previous (not seen)"
    ws_m["B4"] = sum(1 for r in sorted_rows if not r.get("refreshed"))
    ws_m["A5"] = "Paid"
    ws_m["B5"] = sum(1 for r in sorted_rows if (r.get("cost") or 0) > 0)
    ws_m["A6"] = "Free"
    ws_m["B6"] = sum(1 for r in sorted_rows if (r.get("cost") or 0) <= 0)
    ws_m["A8"] = "Source"
    ws_m["B8"] = "skool.com/discovery via Skool Downloader update_skool_xlsx.py"

    wb.save(path)
    log(f"✅ Saved {len(sorted_rows):,} → {path}")


class CatalogUpdater:
    def __init__(self, delay: float = 0.28):
        self.delay = delay
        self._pw = None
        self._browser = None
        self._page = None
        self.build_id = None
        self.languages: List[dict] = []
        self.categories: List[dict] = []

    def open(self):
        from playwright.sync_api import sync_playwright

        self._pw = sync_playwright().start()
        self._browser = self._pw.chromium.launch(headless=True)
        ctx = self._browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
            ),
            locale="en-US",
        )
        self._page = ctx.new_page()
        info = None
        last_err = None
        for attempt in range(4):
            try:
                self._page.goto(
                    "https://www.skool.com/discovery",
                    wait_until="domcontentloaded",
                    timeout=90000,
                )
                for _ in range(25):
                    try:
                        has = self._page.evaluate(
                            "() => !!document.getElementById('__NEXT_DATA__')"
                        )
                    except Exception as e:
                        last_err = e
                        self._page.wait_for_timeout(1000)
                        continue
                    if has:
                        break
                    self._page.wait_for_timeout(700)
                info = self._page.evaluate(
                    """() => {
                      const el = document.getElementById('__NEXT_DATA__');
                      if (!el) return null;
                      const d = JSON.parse(el.textContent);
                      const pp = d.props.pageProps || {};
                      return {
                        buildId: d.buildId,
                        languages: pp.languages || [],
                        categories: pp.categories || [],
                      };
                    }"""
                )
                if info and info.get("buildId"):
                    break
            except Exception as e:
                last_err = e
                time.sleep(1.2)
        if not info or not info.get("buildId"):
            raise RuntimeError(f"Không mở được Discovery: {last_err}")
        self.build_id = info["buildId"]
        self.languages = info.get("languages") or []
        self.categories = info.get("categories") or []
        log(
            f"🌐 Discovery OK buildId={self.build_id} "
            f"langs={len(self.languages)} cats={len(self.categories)}"
        )

    def close(self):
        try:
            if self._browser:
                self._browser.close()
        except Exception:
            pass
        try:
            if self._pw:
                self._pw.stop()
        except Exception:
            pass

    def fetch(self, params: dict) -> Tuple[list, int]:
        qs = urlencode({k: v for k, v in params.items() if v not in (None, "")})
        path = f"/_next/data/{self.build_id}/discovery.json?{qs}"
        for _ in range(3):
            try:
                data = self._page.evaluate(
                    """async (path) => {
                      try {
                        const res = await fetch(path, {headers: {'x-nextjs-data': '1'}});
                        if (!res.ok) return {ok:false, status: res.status, groups:[], num:0};
                        const j = await res.json();
                        const pp = j.pageProps || {};
                        return {ok:true, groups: pp.groups||[], num: pp.numGroups||0};
                      } catch(e) {
                        return {ok:false, error:String(e), groups:[], num:0};
                      }
                    }""",
                    path,
                )
                if data and data.get("ok"):
                    return data.get("groups") or [], int(data.get("num") or 0)
                # soft fail
                if data and data.get("status") in (400, 500):
                    return [], 0
            except Exception:
                self._page.wait_for_timeout(800)
        return [], 0

    def scrape_query(
        self,
        params: dict,
        label: str,
        store: Dict[str, dict],
        lang_tag: Optional[str] = None,
        cat_name: str = "",
        max_pages: int = 80,
    ) -> Tuple[int, int]:
        new_n = 0
        upd_n = 0
        page = 1
        total_pages = "?"
        while page <= max_pages:
            q = dict(params)
            q["p"] = page
            if "srt" not in q:
                q["srt"] = "trending"
            groups, total = self.fetch(q)
            if page == 1 and total:
                total_pages = max(1, (int(total) + 29) // 30)
            if not groups:
                break
            for g in groups:
                item = parse_group(g, lang_tag=lang_tag, cat_name=cat_name)
                if not item:
                    continue
                status = merge_item(store, item)
                if status == "new":
                    new_n += 1
                elif status == "updated":
                    upd_n += 1
            sys.stdout.write(
                f"\r  {label[:48]:<48} p{page}/{total_pages} "
                f"+new {new_n} ~upd {upd_n} | store {len(store):,}   "
            )
            sys.stdout.flush()
            if len(groups) < 30:
                break
            page += 1
            time.sleep(self.delay)
        print()
        return new_n, upd_n


def run_update(
    src: Path,
    out: Path,
    delay: float = 0.28,
    quick: bool = False,
) -> Path:
    store = load_existing(src)
    start = len(store)
    today = datetime.now()
    sheet_date = today.strftime("%Y-%m")
    if out is None:
        out = Path.home() / "Downloads" / f"Skool {today.strftime('%d-%m-%Y')}.xlsx"
    checkpoint = out.with_name(out.stem + ".partial.xlsx")

    def checkpoint_save(tag: str = ""):
        try:
            save_excel(list(store.values()), checkpoint, sheet_date=sheet_date)
            log(f"💾 Checkpoint {tag} → {checkpoint.name} ({len(store):,} rows)")
        except Exception as e:
            log(f"⚠ checkpoint fail: {e}")

    up = CatalogUpdater(delay=delay)
    try:
        up.open()
        langs = [L.get("name") for L in up.languages if L.get("name")]
        cats = [
            (c.get("id"), clean_category(c.get("name") or ""))
            for c in up.categories
            if c.get("id")
        ]

        # 1) all languages
        log("\n[1/5] All languages…")
        for lang in langs:
            up.scrape_query(
                {"lang": lang, "srt": "trending"},
                f"lang={lang}",
                store,
                lang_tag=lang,
            )
        checkpoint_save("after-langs")

        # 2) all categories (all langs)
        log("\n[2/5] All categories…")
        for cid, cname in cats:
            up.scrape_query(
                {"c": cid, "srt": "trending"},
                f"cat={cname}",
                store,
                cat_name=cname,
            )
        checkpoint_save("after-cats")

        if not quick:
            # 3) category × language (all langs — key for coverage beyond 1000 cap)
            log("\n[3/5] Category × language…")
            n_combo = 0
            for cid, cname in cats:
                for lang in langs:
                    up.scrape_query(
                        {"c": cid, "lang": lang, "srt": "trending"},
                        f"{cname[:12]}×{lang}",
                        store,
                        lang_tag=lang,
                        cat_name=cname,
                    )
                    n_combo += 1
                    if n_combo % 40 == 0:
                        checkpoint_save(f"cat×lang-{n_combo}")

            # 4) free/paid × languages
            log("\n[4/5] Price free/paid × languages…")
            for pr in ("free", "paid"):
                up.scrape_query({"pr": pr, "srt": "trending"}, f"pr={pr}", store)
                for lang in langs:
                    up.scrape_query(
                        {"pr": pr, "lang": lang, "srt": "trending"},
                        f"pr={pr}&{lang}",
                        store,
                        lang_tag=lang,
                    )
            checkpoint_save("after-price")

            # 5) free/paid × category × big languages
            log("\n[5/5] Price × category × big languages…")
            for pr in ("free", "paid"):
                for cid, cname in cats:
                    up.scrape_query(
                        {"pr": pr, "c": cid, "srt": "trending"},
                        f"pr={pr}&{cname}",
                        store,
                        cat_name=cname,
                    )
                    for lang in BIG_LANGS:
                        if lang not in langs:
                            continue
                        up.scrape_query(
                            {"pr": pr, "c": cid, "lang": lang, "srt": "trending"},
                            f"{pr[:1]}&{cname[:8]}&{lang[:5]}",
                            store,
                            lang_tag=lang,
                            cat_name=cname,
                        )
            checkpoint_save("after-deep")
        else:
            log("\n[quick] Skip deep cat×lang / price matrices")
            for pr in ("free", "paid"):
                up.scrape_query({"pr": pr, "srt": "trending"}, f"pr={pr}", store)
    finally:
        up.close()

    refreshed = sum(1 for v in store.values() if v.get("refreshed"))
    log(
        f"\n📊 Tổng {len(store):,} (bắt đầu {start:,}, "
        f"refreshed {refreshed:,}, stale {len(store)-refreshed:,})"
    )

    save_excel(list(store.values()), out, sheet_date=sheet_date)

    # also overwrite source Latest if src exists and different path
    if src.exists() and src.resolve() != out.resolve():
        try:
            save_excel(list(store.values()), src, sheet_date=sheet_date)
            log(f"📝 Also updated source: {src}")
        except Exception as e:
            log(f"⚠ Không ghi đè source ({e}) — file mới vẫn OK: {out}")

    # remove partial if final ok
    try:
        if checkpoint.exists() and out.exists():
            checkpoint.unlink()
    except Exception:
        pass

    return out


def repair_notes_excel(src: Path, out: Optional[Path] = None) -> Path:
    """Sửa Notes/Cost không khớp trên file Excel sẵn có (không cào mạng)."""
    out = out or src
    store = load_existing(src)
    if not store:
        raise SystemExit(f"Không load được dữ liệu từ {src}")
    fixed = 0
    for row in store.values():
        # Rows loaded as refreshed=False; repair doesn't mean "stale"
        row["refreshed"] = True
        before = str(row.get("notes") or "")
        after = normalize_notes(row.get("cost") or 0, before)
        # strip accidental stale flags when doing pure repair
        after = re.sub(r"(?i)\s*\|\s*not in latest scrape", "", after).strip(" |")
        after = normalize_notes(row.get("cost") or 0, after)
        if after != before:
            row["notes"] = after
            fixed += 1
        else:
            row["notes"] = after
    sheet_date = datetime.now().strftime("%Y-%m")
    save_excel(list(store.values()), out, sheet_date=sheet_date)
    log(f"🔧 Repaired notes on {fixed:,}/{len(store):,} rows → {out}")
    return out


def main():
    ap = argparse.ArgumentParser(description="Update Skool communities Excel")
    ap.add_argument(
        "--src",
        default=str(DEFAULT_SRC),
        help="Excel cũ (default: ~/Downloads/Skool 14-05-2026.xlsx)",
    )
    ap.add_argument(
        "--out",
        default="",
        help="Excel mới (default: ~/Downloads/Skool DD-MM-YYYY.xlsx)",
    )
    ap.add_argument("--delay", type=float, default=0.28)
    ap.add_argument(
        "--quick",
        action="store_true",
        help="Chỉ cào language + category (nhanh hơn, coverage thấp hơn)",
    )
    ap.add_argument(
        "--repair-notes",
        action="store_true",
        help="Chỉ sửa Notes/Cost không khớp trên Excel (không cào Discovery)",
    )
    args = ap.parse_args()
    src = Path(args.src).expanduser()
    out = Path(args.out).expanduser() if args.out else (
        Path.home() / "Downloads" / f"Skool {datetime.now().strftime('%d-%m-%Y')}.xlsx"
    )
    if args.repair_notes:
        # default: ghi đè file nguồn
        target = Path(args.out).expanduser() if args.out else src
        path = repair_notes_excel(src, target)
        log(f"\nDone → {path}")
        return
    path = run_update(src, out, delay=args.delay, quick=args.quick)
    log(f"\nDone → {path}")


if __name__ == "__main__":
    main()
