#!/usr/bin/env python3
"""
Pass 2: refresh Members/price via Playwright APIRequest (cookies after Discovery).

Resumable flags:
  --skip-refreshed   skip rows already member-refreshed (no stale flag, notes OK)
  --fix-notes-only  only rewrite Notes/Cost consistency (no network)
  --only-inconsistent  only re-fetch rows where Cost>0 but Notes==Free (or similar)
"""
from __future__ import annotations

import argparse
import re
import sys
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
from update_skool_xlsx import (  # noqa: E402
    normalize_cost_raw,
    normalize_notes,
    parse_cost_display,
    repair_notes_excel,
    save_excel,
    slug_from_url,
)

UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
)


def _is_inconsistent(cost: float, notes: str) -> bool:
    n = (notes or "").strip().lower()
    if cost > 0 and (n == "free" or n.startswith("free |") or n == "freemium"):
        return True
    return False


def load_rows(path: Path):
    wb = load_workbook(path, data_only=True)
    ws = wb["Latest"] if "Latest" in wb.sheetnames else wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        d = dict(zip(headers, row))
        url = str(d.get("URL") or "").strip()
        slug = slug_from_url(url)
        if not slug:
            continue
        members_raw = d.get("Members", 0) or 0
        try:
            members = int(float(str(members_raw).replace(",", "")))
        except Exception:
            members = 0
        notes_raw = str(d.get("Notes") or "")
        cost = normalize_cost_raw(d.get("Cost / Month"))
        stale = "not in latest scrape" in notes_raw.lower()
        inconsistent = _is_inconsistent(cost, notes_raw)
        # Member-pass "done" heuristic: not stale and notes match cost
        already = (not stale) and (not inconsistent)
        # Strip stale marker for storage; re-applied only if still not refreshed later
        notes = (
            notes_raw.replace("| not in latest scrape", "")
            .replace("not in latest scrape", "")
            .strip(" |")
        )
        notes = normalize_notes(cost, notes)
        rows.append(
            {
                "slug": slug,
                "name": str(d.get("Name") or ""),
                "tagline": str(d.get("Tagline") or ""),
                "members": members,
                "cost": cost,
                "notes": notes,
                "lang": str(d.get("Language") or ""),
                "category": str(d.get("Category") or ""),
                "url": url or f"https://www.skool.com/{slug}/about",
                "refreshed": already,
                "stale": stale,
                "inconsistent": inconsistent,
            }
        )
    wb.close()
    return rows


def _group_from_props(pp: dict):
    if not isinstance(pp, dict):
        return None
    g = pp.get("currentGroup") or pp.get("group")
    if isinstance(g, dict) and g.get("metadata"):
        return g
    return None


def fetch_about(request, build: str, slug: str, timeout_ms: int = 15000, _depth: int = 0):
    """
    Returns:
      dict with name/members/... on success
      {"missing": True} if 404 / not found
      None on hard error
    """
    if _depth > 3:
        return {"missing": True}
    url = f"https://www.skool.com/_next/data/{build}/{slug}/about.json"
    try:
        r = request.get(url, headers={"x-nextjs-data": "1"}, timeout=timeout_ms)
        if r.status == 404:
            return {"missing": True}
        if not r.ok:
            return None
        j = r.json()
        pp = j.get("pageProps") or {}
        # Next.js client redirect payload
        redir = pp.get("__N_REDIRECT")
        if redir and isinstance(redir, str):
            # e.g. /new-slug/about
            m = re.match(r"^/([^/?#]+)", redir)
            if m and m.group(1) != slug:
                return fetch_about(
                    request, build, m.group(1), timeout_ms=timeout_ms, _depth=_depth + 1
                )
            return {"missing": True}
        g = _group_from_props(pp)
        if not g:
            # try without /about
            url2 = f"https://www.skool.com/_next/data/{build}/{slug}.json"
            r2 = request.get(url2, headers={"x-nextjs-data": "1"}, timeout=timeout_ms)
            if r2.status == 404:
                return {"missing": True}
            if r2.ok:
                j2 = r2.json()
                pp2 = j2.get("pageProps") or {}
                redir2 = pp2.get("__N_REDIRECT")
                if redir2 and isinstance(redir2, str):
                    m2 = re.match(r"^/([^/?#]+)", redir2)
                    if m2 and m2.group(1) != slug:
                        return fetch_about(
                            request,
                            build,
                            m2.group(1),
                            timeout_ms=timeout_ms,
                            _depth=_depth + 1,
                        )
                g = _group_from_props(pp2)
        if not g or not g.get("metadata"):
            return {"missing": True}
        md = g["metadata"]
        members = md.get("totalMembers")
        if members is None:
            members = md.get("memberCount") or 0
        # Pass full price fields so parse_cost_display can use yearly/one-time too
        return {
            "name": md.get("displayName") or g.get("name"),
            "desc": md.get("description") or "",
            "members": int(members or 0),
            "displayPrice": md.get("displayPrice"),
            "currentMBp": md.get("currentMBp"),
            "currentABp": md.get("currentABp"),
            "currentOtBp": md.get("currentOtBp"),
            "model": md.get("membershipModel"),
            "slug": g.get("name") or slug,
        }
    except Exception:
        return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--src",
        default=str(Path.home() / "Downloads" / "Skool 13-07-2026.xlsx"),
    )
    ap.add_argument(
        "--out",
        default=str(Path.home() / "Downloads" / "Skool 13-07-2026.xlsx"),
    )
    ap.add_argument("--start", type=int, default=0)
    ap.add_argument(
        "--skip-refreshed",
        action="store_true",
        help="Bỏ qua row đã refresh (không stale, notes khớp cost)",
    )
    ap.add_argument(
        "--only-inconsistent",
        action="store_true",
        help="Chỉ re-fetch row Cost>0 mà Notes=Free (nhanh, sửa data hỏng)",
    )
    ap.add_argument(
        "--fix-notes-only",
        action="store_true",
        help="Chỉ normalize Notes/Cost trên file, không gọi mạng",
    )
    ap.add_argument("--delay", type=float, default=0.08)
    ap.add_argument("--timeout-ms", type=int, default=15000)
    args = ap.parse_args()

    src = Path(args.src).expanduser()
    if not src.exists():
        alt = Path.home() / "Downloads" / "Skool 13-07-2026.refreshing.xlsx"
        if alt.exists():
            src = alt
        else:
            src = Path.home() / "Downloads" / "Skool 14-05-2026.xlsx"
    out = Path(args.out).expanduser()
    partial = out.with_name(out.stem + ".refreshing.xlsx")
    also = Path.home() / "Downloads" / "Skool 14-05-2026.xlsx"

    if args.fix_notes_only:
        repair_notes_excel(src, out)
        if also.resolve() != out.resolve() and also.exists():
            try:
                repair_notes_excel(out if out.exists() else src, also)
            except Exception as e:
                print(f"warn also-repair: {e}", flush=True)
        return

    rows = load_rows(src)
    by_slug = {r["slug"]: r for r in rows}
    slugs = list(by_slug.keys())
    print(f"Loaded {len(slugs)} from {src.name}", flush=True)

    if args.only_inconsistent:
        todo = [s for s in slugs if by_slug[s].get("inconsistent")]
    elif args.skip_refreshed:
        todo = [s for s in slugs if not by_slug[s].get("refreshed")]
    else:
        todo = slugs[max(0, args.start) :]
    print(f"To refresh: {len(todo)}", flush=True)

    if not todo:
        # still rewrite sanitized notes
        for r in by_slug.values():
            r["notes"] = normalize_notes(r.get("cost") or 0, r.get("notes"))
            r["refreshed"] = True
        sheet_date = datetime.now().strftime("%Y-%m")
        save_excel(list(by_slug.values()), out, sheet_date=sheet_date)
        print(f"Nothing to fetch; saved sanitized → {out}", flush=True)
        return

    ok = fail = missing = 0
    sheet_date = datetime.now().strftime("%Y-%m")
    already = len(slugs) - len(todo)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx = browser.new_context(user_agent=UA, locale="en-US")
        page = ctx.new_page()
        page.goto(
            "https://www.skool.com/discovery",
            wait_until="domcontentloaded",
            timeout=90000,
        )
        page.wait_for_timeout(2000)
        for _ in range(20):
            try:
                if page.evaluate("() => !!document.getElementById('__NEXT_DATA__')"):
                    break
            except Exception:
                pass
            page.wait_for_timeout(700)
        build = page.evaluate(
            "() => JSON.parse(document.getElementById('__NEXT_DATA__').textContent).buildId"
        )
        print(f"buildId {build}", flush=True)
        req = ctx.request

        t0 = time.time()
        for n, slug in enumerate(todo, 1):
            info = fetch_about(req, build, slug, timeout_ms=args.timeout_ms)
            row = by_slug[slug]
            if not info:
                fail += 1
                # keep prior data; do not mark refreshed so resume retries
                row["refreshed"] = False
            elif info.get("missing"):
                missing += 1
                # community removed / renamed — keep last known members, mark done
                row["notes"] = "Missing on Skool"
                row["refreshed"] = True
            else:
                meta = {
                    "displayPrice": info.get("displayPrice"),
                    "currentMBp": info.get("currentMBp"),
                    "currentABp": info.get("currentABp"),
                    "currentOtBp": info.get("currentOtBp"),
                    "membershipModel": info.get("model"),
                }
                cost, notes = parse_cost_display(meta)
                notes = normalize_notes(cost, notes)
                row["name"] = info.get("name") or row["name"]
                desc = re.sub(r"\s+", " ", (info.get("desc") or "").strip())[:300]
                if desc:
                    row["tagline"] = desc
                try:
                    row["members"] = int(info.get("members") or 0)
                except Exception:
                    pass
                row["cost"] = cost
                row["notes"] = notes
                if info.get("slug") and info["slug"] != slug:
                    row["url"] = f"https://www.skool.com/{info['slug']}/about"
                row["refreshed"] = True
                ok += 1

            if n % 25 == 0 or n == len(todo) or n <= 3:
                elapsed = time.time() - t0
                rate = n / elapsed if elapsed else 0
                eta = (len(todo) - n) / rate / 60 if rate else 0
                print(
                    f"  {already + n}/{len(slugs)} ok={ok} miss={missing} fail={fail} "
                    f"rate={rate:.1f}/s eta={eta:.1f}m",
                    flush=True,
                )
            if n % 300 == 0:
                save_excel(list(by_slug.values()), partial, sheet_date=sheet_date)
                print(f"  💾 checkpoint {partial.name}", flush=True)
            if args.delay:
                time.sleep(args.delay)

            # re-open session occasionally if many fails in a row
            if n % 800 == 0:
                try:
                    page.goto(
                        "https://www.skool.com/discovery",
                        wait_until="domcontentloaded",
                        timeout=60000,
                    )
                    page.wait_for_timeout(800)
                    build = page.evaluate(
                        "() => JSON.parse(document.getElementById('__NEXT_DATA__').textContent).buildId"
                    )
                except Exception as e:
                    print(f"  rebind warn: {e}", flush=True)

        browser.close()

    # Sanitize all rows before final write
    for r in by_slug.values():
        r["notes"] = normalize_notes(r.get("cost") or 0, r.get("notes"))
        if r.get("refreshed") is None:
            r["refreshed"] = True

    save_excel(list(by_slug.values()), out, sheet_date=sheet_date)
    try:
        save_excel(list(by_slug.values()), also, sheet_date=sheet_date)
        print(f"Also saved {also.name}", flush=True)
    except Exception as e:
        print(f"warn also-save: {e}", flush=True)

    ref = sum(1 for r in by_slug.values() if r.get("refreshed"))
    print(
        f"DONE ok={ok} miss={missing} fail={fail} total={len(by_slug)} refreshed={ref}",
        flush=True,
    )
    print(f"OUT {out}", flush=True)
    top = sorted(
        by_slug.values(), key=lambda x: x.get("members") or 0, reverse=True
    )[:8]
    for r in top:
        print(
            f"  {r['members']:>8}  {(r.get('name') or '')[:42]}  ${r.get('cost')}  {r.get('notes')}",
            flush=True,
        )


if __name__ == "__main__":
    main()
